from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import pandas as pd
from wcwidth import wcswidth
from openpyxl import load_workbook
import os
import time
import sys

abs_path = os.path.abspath(__file__)
dirt = os.path.dirname(abs_path) + '/'

save_time =time.strftime('%Y%m%d_%H%M',time.localtime(time.time())) 

options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.add_argument('--mute-audio')  # 브라우저에 음소거 옵션을 적용합니다.
options.add_argument('incognito')  # 시크릿 모드의 브라우저가 실행됩니다.
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

key_to_search = input('검색을 원하는 종명 혹은 속명을 영어로 입력하시오 : ')


driver.get(
    'https://species.nibr.go.kr/home/mainHome.do?searchType=total&cont_link=009&subMenu=009001&contCd=009001&searchField=')
driver.implicitly_wait(5)
driver.find_element(By.XPATH, '//*[@id="div_search_sub_btn"]/a').click()
driver.implicitly_wait(5)
search_box = driver.find_element(By.XPATH, '//*[@id="topSearchKeyword"]')
search_box.send_keys(key_to_search)
search_box.send_keys(Keys.RETURN)
#//*[@id="query"]
driver.switch_to.frame('gwpMain')

driver.implicitly_wait(5)
(driver.find_element(By.XPATH, '//*[@id="contents"]/div[3]/ul/li[6]/a')).click() #표본 click


try:
    WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="plusTab_lifeSpecimen"]/div/div[1]/span[2]')))
except:
    sys.exit("검색 결과 없음")

total_species_num = (driver.find_element(By.XPATH,'//*[@id="plusTab_lifeSpecimen"]/div/div[1]/span[2]').text)
sp_num_re = '\d{1,10}'
sp_num_result = re.search(sp_num_re, total_species_num)

total_sp = int(((sp_num_result.group()).strip()))

nibr_dict = {}

def split_species(n):
    a = driver.find_element(By.XPATH, '//*[@id="plusTab_lifeSpecimen"]/ul/li[%s]/p[1]' % n).text[:-6] #표본번호, 표본유형
    b = driver.find_element(By.XPATH, '//*[@id="plusTab_lifeSpecimen"]/ul/li[%s]/p[2]' % n).text #생물 (국명, 학명)
    c = driver.find_element(By.XPATH, '//*[@id="plusTab_lifeSpecimen"]/ul/li[%s]/p[3]' % n).text #채집지, 채집일

    specimen_num_re = '표본번호 : ([A-Za-z0-9]+)'
    specimen_state_re = '표본유형 : (\S+)'
    kor_re = '생물\(국명, 학명\) : ([가-힣]+)'
    eng_re = key_to_search + '(\s)?(aff\. )?(cf\. )?([a-z]+)?(sp\.)?( var\. [a-z]+)?'

    kor_pt = re.compile(r'[ㄱ-ㅣ가-힣]')
    
    specimen_num_result = re.search(specimen_num_re, a)
    specimen_state_result = re.search(specimen_state_re, a)
    kor_result = re.search(kor_re, b)
    eng_result = re.search(eng_re, b)


    if kor_result:
        kor_name = (kor_result.group(1)).strip()
    else:
        kor_name = '국명정보 없음'

    
    if eng_result:
        scientific_name = (eng_result.group()).strip()
    else :
        sys.exit("입력을 확인하시오")

    if specimen_state_result:
        specimen_state = (specimen_state_result.group(1)).strip()
    else:
        specimen_state = '미상'

    if '|' in c:
        c_1 = c.split('|')
        c_1.remove('')

        if len(c_1) > 1:
            habitat = (c_1[0].strip('채집지 : ')).strip()
            col_date = ((c_1[1].strip('채집일 : ')).rstrip('|')).strip()
            
        else:
            if '채집지 : ' in c: 
                habitat = (c.strip('채집지 : ')).strip()
                col_date = '미상'

            if '채집일 : ' in c:
                habitat = '미상'
                col_date = ((c.strip('채집일 : ')).rstrip('|')).strip()

    if col_date == '0000':
        col_date = '미상'
                

    result =re.match(kor_pt,habitat)
    specimen_num = (specimen_num_result.group(1)).strip()

    if result :
        col_place = '국내'        
    else:
        col_place = '해외'

        
    mid_list = [ specimen_num, scientific_name,kor_name,specimen_state,col_place,habitat,col_date,'','','','']
    return specimen_num, mid_list

try: #표본 갯수가 10개 이상일 경우
    page_re = '(\d{1,})'
    page_info = driver.find_element(By.XPATH,'//*[@id="pagewrap"]/span').text
    page_list = re.findall(page_re,page_info)

    page_s = int(page_list[0])
    page_e = int(page_list[1])
    while page_s <= page_e:

        num = driver.find_elements(By.XPATH, '//*[@id="plusTab_lifeSpecimen"]/ul/li')#각 페이지에 있는 표본의 갯수
        for n in range(1, len(num) + 1):
            specimen_num, mid_list = split_species(n)
            nibr_dict[specimen_num] = mid_list
        
        driver.execute_script(f"javascript:doPaging('{page_s*10}');")
        page_s +=1
except: #표본갯수가 10개 미만일 경우
    num = driver.find_elements(By.XPATH, '//*[@id="plusTab_lifeSpecimen"]/ul/li')#각 페이지에 있는 표본의 갯수
    for n in range(1, len(num) + 1):
        specimen_num, mid_list = split_species(n)
        nibr_dict[specimen_num] = mid_list

print('정보수집 완료')

#def crawl_seq():
    #result_dict = collect_info()
    
result_dict = nibr_dict
print('sequenc정보 수집 시작')

driver.find_element(By.XPATH,'//*[@id="contents"]/div[3]/ul/li[8]/a').click()

driver.implicitly_wait(5)

try:        
    info = driver.find_element(By.XPATH,'//*[@id="plusTab_lifeGene"]/ul/li/p').text
    
    nn ='\d'
    seq_num = (driver.find_element(By.XPATH,'//*[@id="plusTab_lifeGene"]/div/div[1]/span[2]')).text
    seq_num_mo = (re.search(nn,seq_num)).group()
    
    a =(driver.find_elements(By.XPATH,'//*[@id="plusTab_lifeGene"]/ul/li'))#페이지에 몇개 있는가 정보
    
    if int(seq_num_mo) == 1:
        driver.find_element(By.XPATH,'//*[@id="plusTab_lifeGene"]/ul/li/p/a').click()
        
        driver.switch_to.window(driver.window_handles[-1])
        
        driver.find_element(By.XPATH,'//*[@id="check_lineNum"]').click()
        
        sample_ID = driver.find_element(By.XPATH,'//*[@id="popupwrap"]/div[2]/div[2]/table/tbody/tr[1]/td[2]').text
        marker_name = driver.find_element(By.XPATH,'//*[@id="popupwrap"]/div[2]/div[2]/table/tbody/tr[4]/td[1]').text
        sample_seq = driver.find_element(By.XPATH,'//*[@id="strSeq"]').text
        sample_seq =(sample_seq.replace(' ','')).replace('\n','')
        
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
        if marker_name == 'ITS':
            ((result_dict[sample_ID])[7])=sample_seq
            
        if marker_name == 'LSU':
            ((result_dict[sample_ID])[8])=sample_seq
            
        if marker_name == 'RPB':
            ((result_dict[sample_ID])[9])=sample_seq
            
        if marker_name == 'EF':
            ((result_dict[sample_ID])[10])=sample_seq
                
    elif int(seq_num)>1:    
        for i in range(1,seq_num+1):
            driver.find_element(By.XPATH,f'//*[@id="plusTab_lifeGene"]/ul/li[{i}]/p[1]/a').click()
        
            driver.switch_to.window(driver.window_handles[-1])
            
            driver.find_element(By.XPATH,'//*[@id="check_lineNum"]').click()
            
            sample_ID = driver.find_element(By.XPATH,'//*[@id="popupwrap"]/div[2]/div[2]/table/tbody/tr[1]/td[2]').text
            marker_name = driver.find_element(By.XPATH,'//*[@id="popupwrap"]/div[2]/div[2]/table/tbody/tr[4]/td[1]').text
            sample_seq = driver.find_element(By.XPATH,'//*[@id="strSeq"]').text
            sample_seq =(sample_seq.replace(' ','')).replace('\n','')
    
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
except:
    print('sequence 정보 없음')
    pass

driver.close()

print('저장 시작')
nibr_list = list(result_dict.values())

save_time =time.strftime('%Y%m%d_%H%M',time.localtime(time.time())) 

nibr_df = pd.DataFrame.from_records(nibr_list)
excel_dir = f'{dirt}{key_to_search}_nibr_result({save_time}).xlsx'

nibr_df.columns =['spcimen_no', 'scientific_name','kor_name','specimen_type','col_place','habitat','col_date','ITS','LSU','RPB','EF']

nibr_df =nibr_df.sort_values(by=['scientific_name','col_date'], ascending=[True,True])
with pd.ExcelWriter(excel_dir) as writer:
    nibr_df.to_excel(writer, sheet_name ='nibr', index=False)


wb = load_workbook(filename=excel_dir, data_only=True)
ws = wb.active
col_range = ws['A:G']
for column_cells in col_range:
    length = max(wcswidth(str(cell.value)) * 1.1 for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length
wb.save(excel_dir)

print('xlsx 파일 저장됨')


